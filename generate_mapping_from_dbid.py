import csv
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required. Install with: pip install openpyxl"
    ) from exc


SOURCE_CSV = Path("dropbox_full_metadata.csv")
DBID_XLSX = Path("dbid_mapping.xlsx")
OUTPUT_CSV = Path("mapping_full_metadata.csv")
FALLBACK = "sharepoint.admin@ducorpgroup.com"
LIBRARY_PREFIX = "/sites/Testing-site/Shared Documents"



def load_dbid_mapping():
    if not DBID_XLSX.exists():
        raise FileNotFoundError(f"Mapping workbook not found: {DBID_XLSX}")

    workbook = load_workbook(DBID_XLSX, read_only=True, data_only=True)
    sheet = workbook.active

    headers = {str(cell.value).strip().lower(): idx for idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1)), start=0)}

    required = {"dbid", "email"}
    if not required.issubset(headers.keys()):
        missing = required - headers.keys()
        raise ValueError(f"Missing required columns in {DBID_XLSX}: {', '.join(missing)}")

    dbid_to_email = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        dbid_val = str(row[headers["dbid"]] or "").strip()
        email_val = str(row[headers["email"]] or "").strip()
        if not dbid_val:
            continue
        if not email_val:
            email_val = FALLBACK
        dbid_to_email[dbid_val] = email_val

    return dbid_to_email


def build_mapping(dbid_to_email):
    if not SOURCE_CSV.exists():
        raise FileNotFoundError(f"Source CSV not found: {SOURCE_CSV}")

    with SOURCE_CSV.open(newline="", encoding="utf-8") as src, OUTPUT_CSV.open(
        "w", newline="", encoding="utf-8"
    ) as dst:
        reader = csv.DictReader(src)
        fieldnames = [
            "FilePath",
            "AuthorUPN",
            "EditorUPN",
            "created_by_id",
            "created_by_email",
            "created_by_name",
            "last_modified_by_id",
            "last_modified_by_email",
            "last_modified_by_name",
            "created_client_modified",
            "created_server_modified",
            "created_source",
            "last_client_modified",
            "last_server_modified",
            "rev",
            "size_mb",
        ]
        writer = csv.DictWriter(dst, fieldnames=fieldnames)
        writer.writeheader()

        rows_written = 0
        for row in reader:
            if row.get("type") != "file":
                continue

            path = (row.get("path") or "").strip()
            if not path:
                continue

            created_id = (row.get("created_by_id") or "").strip()
            modified_id = (row.get("last_modified_by_id") or "").strip()

            author = dbid_to_email.get(created_id, (row.get("created_by_email") or "").strip() or FALLBACK)
            editor = dbid_to_email.get(modified_id, (row.get("last_modified_by_email") or "").strip() or author)

            writer.writerow(
                {
                    "FilePath": LIBRARY_PREFIX.rstrip("/") + path,
                    "AuthorUPN": author,
                    "EditorUPN": editor,
                    "created_by_id": created_id,
                    "created_by_email": (row.get("created_by_email") or "").strip(),
                    "created_by_name": (row.get("created_by_name") or "").strip(),
                    "last_modified_by_id": modified_id,
                    "last_modified_by_email": (row.get("last_modified_by_email") or "").strip(),
                    "last_modified_by_name": (row.get("last_modified_by_name") or "").strip(),
                    "created_client_modified": (row.get("created_client_modified") or "").strip(),
                    "created_server_modified": (row.get("created_server_modified") or "").strip(),
                    "created_source": (row.get("created_source") or "").strip(),
                    "last_client_modified": (row.get("last_client_modified") or "").strip(),
                    "last_server_modified": (row.get("last_server_modified") or "").strip(),
                    "rev": (row.get("rev") or "").strip(),
                    "size_mb": (row.get("size_mb") or "").strip(),
                }
            )
            rows_written += 1

    print(f"✅ Wrote {rows_written} rows to {OUTPUT_CSV}")


def main():
    dbid_to_email = load_dbid_mapping()
    build_mapping(dbid_to_email)


if __name__ == "__main__":
    main()

